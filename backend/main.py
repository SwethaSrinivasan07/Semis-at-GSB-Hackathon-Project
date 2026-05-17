"""
FastAPI app for SupplyLine backend.

Wraps the same backend modules the Streamlit demo uses, so a future
React/Next.js frontend can call the API without touching the Python.

Run locally:
    uvicorn backend.main:app --reload --port 8000

Endpoints:
    GET  /health           liveness check
    GET  /part/{mpn}       distributor data for a single MPN
    POST /analyze          upload a BOM file (csv | xlsx | json), get analysis JSON
"""

from __future__ import annotations

import csv
import io
import json
import logging
from typing import Any

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware

from . import demo_mode, distributor_api

log = logging.getLogger(__name__)

app = FastAPI(title="SupplyLine API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # tighten to specific frontend origins before prod
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/part/{mpn}")
def get_part(mpn: str) -> dict[str, Any]:
    record = distributor_api.search_part(mpn)
    if record is None:
        raise HTTPException(status_code=404, detail=f"MPN {mpn!r} not found in any source")
    return record


@app.post("/analyze")
async def analyze(file: UploadFile = File(...)) -> dict[str, Any]:
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    name = (file.filename or "").lower()
    try:
        if name.endswith(".json"):
            bom_lines = _parse_json_bom(raw)
        elif name.endswith(".csv"):
            bom_lines = _parse_csv_bom(raw)
        elif name.endswith((".xlsx", ".xls")):
            bom_lines = _parse_xlsx_bom(raw)
        else:
            raise HTTPException(status_code=415, detail="Unsupported file type. Use .csv, .xlsx, or .json")
    except HTTPException:
        raise
    except Exception as exc:
        log.exception("BOM parse failed")
        raise HTTPException(status_code=400, detail=f"Could not parse BOM: {exc}")

    if not bom_lines:
        raise HTTPException(status_code=400, detail="No BOM lines extracted from file")

    return demo_mode.analyze_bom(bom_lines, bom_name=file.filename or "uploaded BOM")


# ---------- BOM parsers (lightweight; bom_parser.py supersedes if present) ----

def _parse_json_bom(raw: bytes) -> list[dict[str, Any]]:
    payload = json.loads(raw.decode("utf-8"))
    lines = payload.get("lines", payload) if isinstance(payload, dict) else payload
    return [_normalize_line(item) for item in lines]


def _parse_csv_bom(raw: bytes) -> list[dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(raw.decode("utf-8")))
    return [_normalize_line(row) for row in reader]


def _parse_xlsx_bom(raw: bytes) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        row_dict = {headers[i]: r[i] for i in range(len(headers))}
        out.append(_normalize_line(row_dict))
    return out


def _normalize_line(row: dict[str, Any]) -> dict[str, Any]:
    """Map common header variants to the (mpn, qty, ref_designator) contract."""
    keys = {k.lower(): k for k in row.keys() if k}
    def pick(*candidates: str) -> Any:
        for c in candidates:
            if c in keys:
                return row[keys[c]]
        return None
    return {
        "ref_designator": pick("ref_designator", "ref", "reference", "designator"),
        "mpn": pick("mpn", "part_number", "partnumber", "manufacturer_part_number"),
        "qty": int(pick("qty", "quantity") or 1),
    }
