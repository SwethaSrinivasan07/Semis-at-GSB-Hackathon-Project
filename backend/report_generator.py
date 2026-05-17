"""
Report generation — PDF and annotated Excel.

Consumes the analyzed-BOM dict produced by demo_mode.analyze_bom() (and
eventually by the FastAPI /analyze endpoint). Output design choices:

PDF:
- Cover/executive summary on page 1 (counts, top 3 risks, #1 recommendation)
- Per-line risk table with a thick colored *left border* per row (more
  scannable than cell-fill alone)
- A "Flagged Parts — Substitutes" section listing up to 2 alternates per
  RED/YELLOW line
- Page numbers and generation date in the footer

Excel:
- Tab 1 "Risk Summary" — narrative + counts + top risks + recommendation
- Tab 2 "Annotated BOM" — every line, row fill colored by risk level,
  AutoFilter with a Critical Parts preset (RED only), datasheet hyperlink
- Tab 3 "Substitutes" — every substitute recommendation, grouped by the
  flagged part it covers
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn, Filters
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

RISK_FILL = {
    "RED": PatternFill("solid", fgColor="FFC9C9"),
    "YELLOW": PatternFill("solid", fgColor="FFF1B8"),
    "GREEN": PatternFill("solid", fgColor="D6F5D6"),
}

RISK_BORDER = {
    "RED": colors.HexColor("#C0392B"),
    "YELLOW": colors.HexColor("#D4AC0D"),
    "GREEN": colors.HexColor("#1E8449"),
}


# ====================================================================
# Excel
# ====================================================================

def write_excel(analysis: dict[str, Any], path: str | Path) -> Path:
    path = Path(path)
    wb = Workbook()

    _write_risk_summary_tab(wb, analysis)
    _write_annotated_bom_tab(wb, analysis)
    _write_substitutes_tab(wb, analysis)

    # Remove the default sheet created by Workbook() if it's still empty.
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row <= 1 and wb["Sheet"].max_column <= 1:
        del wb["Sheet"]

    wb.save(path)
    return path


def _write_risk_summary_tab(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Risk Summary", 0)
    s = analysis["summary"]

    ws["A1"] = "Catena — Risk Summary"
    ws["A1"].font = Font(bold=True, size=18)
    ws.merge_cells("A1:E1")

    ws["A2"] = f"BOM: {analysis.get('bom_name', '(unnamed)')}"
    ws["A3"] = f"Generated: {analysis.get('generated_at', '')}"

    ws["A5"] = "Total lines"
    ws["B5"] = s["total_lines"]
    ws["A6"] = "RED"
    ws["B6"] = s["red_count"]
    ws["B6"].fill = RISK_FILL["RED"]
    ws["A7"] = "YELLOW"
    ws["B7"] = s["yellow_count"]
    ws["B7"].fill = RISK_FILL["YELLOW"]
    ws["A8"] = "GREEN"
    ws["B8"] = s["green_count"]
    ws["B8"].fill = RISK_FILL["GREEN"]

    ws["A10"] = "Top risks"
    ws["A10"].font = Font(bold=True)
    for i, risk in enumerate(s.get("top_risks") or [], start=11):
        ws.cell(row=i, column=1, value=f"• {risk}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    row = 11 + max(len(s.get("top_risks") or []), 1) + 1
    ws.cell(row=row, column=1, value="Recommendation").font = Font(bold=True)
    ws.cell(row=row + 1, column=1, value=s.get("recommendation", ""))
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=5)
    ws.cell(row=row + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    row += 4
    ws.cell(row=row, column=1, value="Narrative").font = Font(bold=True)
    ws.cell(row=row + 1, column=1, value=analysis.get("narrative", ""))
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=5)
    ws.cell(row=row + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row + 1].height = 80

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 28


def _write_annotated_bom_tab(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Annotated BOM")

    headers = [
        "Ref", "Qty", "MPN", "Manufacturer", "Category", "Lifecycle",
        "Lead (wks)", "Stock", "Fab country", "Supplier count",
        "Risk score", "Risk level", "Risk factors", "Datasheet",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")
        cell.border = Border(bottom=Side(style="thin", color="888888"))

    for line in analysis["lines"]:
        part = line.get("part") or {}
        ws.append([
            line.get("ref_designator"),
            line.get("qty"),
            line.get("mpn"),
            part.get("manufacturer"),
            part.get("category"),
            part.get("lifecycle"),
            part.get("lead_time_weeks"),
            part.get("stock"),
            part.get("fab_country"),
            part.get("supplier_count"),
            line.get("risk_score"),
            line.get("risk_level"),
            "; ".join(line.get("risk_factors") or []),
            part.get("datasheet_url"),
        ])
        last_row = ws.max_row
        fill = RISK_FILL.get(line.get("risk_level", "GREEN"))
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=last_row, column=col_idx).fill = fill
        ds = part.get("datasheet_url")
        if ds:
            cell = ws.cell(row=last_row, column=len(headers))
            cell.hyperlink = ds
            cell.value = ds
            cell.font = Font(color="0563C1", underline="single")

    widths = [6, 6, 30, 18, 14, 12, 11, 10, 14, 14, 11, 12, 50, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_col = get_column_letter(len(headers))
    ref = f"A1:{last_col}{ws.max_row}"
    ws.auto_filter.ref = ref
    # "Critical Parts" preset: filter the Risk level column down to RED.
    risk_level_col_index = headers.index("Risk level")
    filter_col = FilterColumn(colId=risk_level_col_index)
    filter_col.filters = Filters(filter=["RED"])
    ws.auto_filter.filterColumn.append(filter_col)

    ws.freeze_panes = "A2"


def _write_substitutes_tab(wb: Workbook, analysis: dict[str, Any]) -> None:
    ws = wb.create_sheet("Substitutes")
    ws.append([
        "Flagged MPN", "Flagged manufacturer", "Risk level",
        "Substitute MPN", "Substitute manufacturer", "Sub lead (wks)",
        "Sub stock", "Sub fab country", "Sub price USD",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")

    for line in analysis["lines"]:
        subs = line.get("substitutes") or []
        if not subs or line.get("risk_level") == "GREEN":
            continue
        part = line.get("part") or {}
        for sub in subs:
            ws.append([
                line.get("mpn"),
                part.get("manufacturer"),
                line.get("risk_level"),
                sub.get("mpn"),
                sub.get("manufacturer"),
                sub.get("lead_time_weeks"),
                sub.get("stock"),
                sub.get("fab_country"),
                sub.get("price_usd"),
            ])
            fill = RISK_FILL.get(line.get("risk_level"))
            if fill:
                ws.cell(row=ws.max_row, column=3).fill = fill

    widths = [26, 18, 12, 26, 18, 12, 10, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ====================================================================
# PDF
# ====================================================================

class _FooterCanvas(canvas.Canvas):
    """Adds 'Page N of M  |  generated <date>  |  Catena' to every page."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_pages: list[dict] = []

    def showPage(self):
        self._saved_pages.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._saved_pages)
        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        for state in self._saved_pages:
            self.__dict__.update(state)
            self._draw_footer(total, stamp)
            super().showPage()
        super().save()

    def _draw_footer(self, total: int, stamp: str) -> None:
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#666666"))
        self.drawString(0.5 * inch, 0.4 * inch, f"Catena  |  Generated {stamp}")
        self.drawRightString(
            LETTER[0] - 0.5 * inch, 0.4 * inch,
            f"Page {self._pageNumber} of {total}",
        )


def write_pdf(analysis: dict[str, Any], path: str | Path) -> Path:
    path = Path(path)
    doc = BaseDocTemplate(
        str(path),
        pagesize=LETTER,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.7 * inch,
        title="Catena BOM Risk Report",
    )
    frame = Frame(
        doc.leftMargin, doc.bottomMargin,
        doc.width, doc.height,
        id="body",
    )
    doc.addPageTemplates([PageTemplate(id="main", frames=[frame])])

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=18, spaceAfter=6)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=10, leading=13)
    body_small = ParagraphStyle("body_small", parent=body, fontSize=9, leading=11)

    story: list[Any] = []

    # ----- Header -----
    story.append(Paragraph("Catena — BOM Risk Report", h1))
    story.append(Paragraph(
        f"<b>BOM:</b> {analysis.get('bom_name', '(unnamed)')} &nbsp;&nbsp; "
        f"<b>Generated:</b> {analysis.get('generated_at', '')}",
        body,
    ))
    story.append(Spacer(1, 10))

    # ----- Executive summary -----
    s = analysis["summary"]
    story.append(Paragraph("Executive summary", h2))
    bullets = [
        f"<b>{s['red_count']}</b> RED, <b>{s['yellow_count']}</b> YELLOW, "
        f"<b>{s['green_count']}</b> GREEN out of {s['total_lines']} lines analyzed.",
    ]
    if s.get("top_risks"):
        bullets.append("Top risks: " + "; ".join(s["top_risks"]))
    if s.get("recommendation"):
        bullets.append("Recommendation: " + s["recommendation"])
    for b in bullets:
        story.append(Paragraph("• " + b, body))
    story.append(Spacer(1, 8))

    if analysis.get("narrative"):
        story.append(Paragraph("Narrative", h2))
        story.append(Paragraph(analysis["narrative"], body))
        story.append(Spacer(1, 10))

    # ----- Per-line risk table with colored left borders -----
    story.append(Paragraph("Line-by-line risk", h2))
    header = ["Ref", "MPN", "Mfr", "Cat", "LT (wk)", "Stock", "Lifecycle", "Score", "Level"]
    rows: list[list[str]] = [header]
    for line in analysis["lines"]:
        part = line.get("part") or {}
        rows.append([
            str(line.get("ref_designator", "")),
            str(line.get("mpn", "")),
            str(part.get("manufacturer", "")),
            str(part.get("category", "")),
            str(part.get("lead_time_weeks", "")),
            str(part.get("stock", "")),
            str(part.get("lifecycle", "")),
            str(line.get("risk_score", "")),
            str(line.get("risk_level", "")),
        ])
    col_widths = [0.45 * inch, 1.8 * inch, 1.1 * inch, 0.9 * inch,
                  0.55 * inch, 0.55 * inch, 0.7 * inch, 0.5 * inch, 0.55 * inch]
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEEEEE")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
    ]
    # Thick colored left border per data row
    for i, line in enumerate(analysis["lines"], start=1):
        color = RISK_BORDER.get(line.get("risk_level"), colors.grey)
        style_cmds.append(("LINEBEFORE", (0, i), (0, i), 4, color))
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    story.append(Spacer(1, 12))

    # ----- Substitutes section -----
    flagged = [l for l in analysis["lines"] if l.get("risk_level") in ("RED", "YELLOW") and l.get("substitutes")]
    if flagged:
        story.append(Paragraph("Recommended substitutes (top 2 per flagged part)", h2))
        for line in flagged:
            part = line.get("part") or {}
            story.append(Paragraph(
                f"<b>{line.get('mpn')}</b> ({part.get('manufacturer', '?')}) — {line.get('risk_level')}",
                body,
            ))
            for sub in (line.get("substitutes") or [])[:2]:
                story.append(Paragraph(
                    f"&nbsp;&nbsp;&nbsp;→ {sub.get('mpn')} ({sub.get('manufacturer')}) — "
                    f"{sub.get('lead_time_weeks')} wk lead, stock {sub.get('stock')}, "
                    f"fab {sub.get('fab_country')}, ${sub.get('price_usd')}",
                    body_small,
                ))
            story.append(Spacer(1, 4))

    doc.build(story, canvasmaker=_FooterCanvas)
    return path
